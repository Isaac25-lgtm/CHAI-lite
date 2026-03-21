from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, g
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import io
import os
import re
import time
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import secrets

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-this-in-production')

# Use DATABASE_URL env var in production (Neon/Render), SQLite locally as fallback
DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///chai_local.db')

# Fix for older postgres:// URIs (Render/Heroku style)
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'pool_size': 5,
    'max_overflow': 10,
    'connect_args': {'connect_timeout': 5}
}

db = SQLAlchemy(app)

# Admin credentials from environment variables
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD_RAW = os.environ.get('ADMIN_PASSWORD', 'admin123')
ADMIN_PASSWORD_HASH = generate_password_hash(ADMIN_PASSWORD_RAW)

# Rate limiting: track failed login/PIN attempts by IP
# Structure: { ip: [(timestamp, ...), ...] }
_failed_login_attempts = {}
_failed_pin_attempts = {}
RATE_LIMIT_MAX = 5       # max failures
RATE_LIMIT_WINDOW = 900  # 15 minutes in seconds
RATE_LIMIT_BLOCK = 900   # block for 15 minutes


def _clean_attempts(store, ip):
    """Remove attempts older than the window"""
    now = time.time()
    if ip in store:
        store[ip] = [t for t in store[ip] if now - t < RATE_LIMIT_WINDOW]
        if not store[ip]:
            del store[ip]


def _is_rate_limited(store, ip):
    """Check if IP is rate-limited"""
    _clean_attempts(store, ip)
    if ip not in store:
        return False
    return len(store[ip]) >= RATE_LIMIT_MAX


def _record_failure(store, ip):
    """Record a failed attempt"""
    now = time.time()
    if ip not in store:
        store[ip] = []
    store[ip].append(now)


# Session timeout (30 minutes)
SESSION_TIMEOUT = int(os.environ.get('SESSION_TIMEOUT_MINUTES', 30))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=SESSION_TIMEOUT)


@app.before_request
def enforce_https():
    """Redirect HTTP to HTTPS when behind a proxy (Render) - only in production"""
    if not app.debug and request.headers.get('X-Forwarded-Proto') == 'http':
        url = request.url.replace('http://', 'https://', 1)
        return redirect(url, code=301)


@app.before_request
def check_session_timeout():
    """Auto-logout after inactivity"""
    if 'admin_logged_in' in session:
        last_active = session.get('last_active')
        if last_active:
            last_dt = datetime.fromisoformat(last_active)
            if datetime.utcnow() - last_dt > timedelta(minutes=SESSION_TIMEOUT):
                session.clear()
                flash('Session expired due to inactivity. Please log in again.', 'warning')
                return redirect(url_for('admin_login'))
        session['last_active'] = datetime.utcnow().isoformat()


def ordinal(n):
    ords = ['', '1st', '2nd', '3rd', '4th', '5th', '6th', '7th']
    return ords[n] if n < len(ords) else f'{n}th'


# ── Models ──────────────────────────────────────────────────────────

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assessment_id = db.Column(db.Integer, nullable=True)
    action = db.Column(db.String(50), nullable=False)  # create, edit, delete, lock, unlock, clear
    entity_type = db.Column(db.String(50), nullable=False)  # registration, bank_detail, assessment
    entity_id = db.Column(db.Integer, nullable=True)
    details = db.Column(db.Text, default='')
    performed_by = db.Column(db.String(100), default='system')
    performed_at = db.Column(db.DateTime, default=datetime.utcnow)


def log_audit(assessment_id, action, entity_type, entity_id=None, details=''):
    """Log an audit trail entry"""
    try:
        entry = AuditLog(
            assessment_id=assessment_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            details=details,
            performed_by=session.get('admin_username', 'field_user')
        )
        db.session.add(entry)
        db.session.commit()
    except Exception:
        db.session.rollback()


class Assessment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    start_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)
    campaign_days = db.Column(db.Integer, default=3)
    pin = db.Column(db.String(10), nullable=False)
    created_by = db.Column(db.String(100), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    is_locked = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    registrations = db.relationship('Registration', backref='assessment', lazy=True)
    bank_details = db.relationship('BankDetail', backref='assessment', lazy=True)

    @property
    def dates_label(self):
        """Format dates nicely for display"""
        if self.start_date and self.end_date:
            return f"{self.start_date.strftime('%d %b %Y')} to {self.end_date.strftime('%d %b %Y')}"
        elif self.start_date:
            return f"From {self.start_date.strftime('%d %b %Y')}"
        return ''

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'dates_label': self.dates_label,
            'start_date': self.start_date.strftime('%Y-%m-%d') if self.start_date else '',
            'end_date': self.end_date.strftime('%Y-%m-%d') if self.end_date else '',
            'campaign_days': self.campaign_days,
            'pin': self.pin,
            'is_active': self.is_active,
        }


class BankDetail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assessment_id = db.Column(db.Integer, db.ForeignKey('assessment.id'), nullable=False)
    participant_name = db.Column(db.String(200), default='')
    account_name = db.Column(db.String(200), nullable=False)
    designation = db.Column(db.String(100), default='')
    bank_name = db.Column(db.String(100), nullable=False)
    account_number = db.Column(db.String(50), nullable=False)
    branch = db.Column(db.String(100), default='')
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Campaign day attendance (same as registration)
    day1 = db.Column(db.Boolean, default=False)
    day2 = db.Column(db.Boolean, default=False)
    day3 = db.Column(db.Boolean, default=False)
    day4 = db.Column(db.Boolean, default=False)
    day5 = db.Column(db.Boolean, default=False)
    day6 = db.Column(db.Boolean, default=False)
    day7 = db.Column(db.Boolean, default=False)
    day8 = db.Column(db.Boolean, default=False)
    day9 = db.Column(db.Boolean, default=False)
    day10 = db.Column(db.Boolean, default=False)
    day11 = db.Column(db.Boolean, default=False)
    day12 = db.Column(db.Boolean, default=False)
    day13 = db.Column(db.Boolean, default=False)
    day14 = db.Column(db.Boolean, default=False)
    day15 = db.Column(db.Boolean, default=False)
    day16 = db.Column(db.Boolean, default=False)
    day17 = db.Column(db.Boolean, default=False)
    day18 = db.Column(db.Boolean, default=False)
    day19 = db.Column(db.Boolean, default=False)
    day20 = db.Column(db.Boolean, default=False)
    day21 = db.Column(db.Boolean, default=False)
    day22 = db.Column(db.Boolean, default=False)
    day23 = db.Column(db.Boolean, default=False)
    day24 = db.Column(db.Boolean, default=False)
    day25 = db.Column(db.Boolean, default=False)
    day26 = db.Column(db.Boolean, default=False)
    day27 = db.Column(db.Boolean, default=False)
    day28 = db.Column(db.Boolean, default=False)
    day29 = db.Column(db.Boolean, default=False)
    day30 = db.Column(db.Boolean, default=False)

    def to_dict(self):
        d = {
            'id': self.id,
            'assessment_id': self.assessment_id,
            'participant_name': self.participant_name or '',
            'account_name': self.account_name,
            'designation': self.designation,
            'bank_name': self.bank_name,
            'account_number': self.account_number,
            'branch': self.branch,
            'submitted_at': self.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
        }
        for i in range(1, 31):
            d[f'day{i}'] = getattr(self, f'day{i}', False)
        return d


class Registration(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assessment_id = db.Column(db.Integer, db.ForeignKey('assessment.id'), nullable=False)
    participant_name = db.Column(db.String(200), nullable=False)
    cadre = db.Column(db.String(100), nullable=False)
    district = db.Column(db.String(100), nullable=False)
    facility = db.Column(db.String(200), nullable=False)
    registration_date = db.Column(db.Date, nullable=False)
    day1 = db.Column(db.Boolean, default=False)
    day2 = db.Column(db.Boolean, default=False)
    day3 = db.Column(db.Boolean, default=False)
    day4 = db.Column(db.Boolean, default=False)
    day5 = db.Column(db.Boolean, default=False)
    day6 = db.Column(db.Boolean, default=False)
    day7 = db.Column(db.Boolean, default=False)
    day8 = db.Column(db.Boolean, default=False)
    day9 = db.Column(db.Boolean, default=False)
    day10 = db.Column(db.Boolean, default=False)
    day11 = db.Column(db.Boolean, default=False)
    day12 = db.Column(db.Boolean, default=False)
    day13 = db.Column(db.Boolean, default=False)
    day14 = db.Column(db.Boolean, default=False)
    day15 = db.Column(db.Boolean, default=False)
    day16 = db.Column(db.Boolean, default=False)
    day17 = db.Column(db.Boolean, default=False)
    day18 = db.Column(db.Boolean, default=False)
    day19 = db.Column(db.Boolean, default=False)
    day20 = db.Column(db.Boolean, default=False)
    day21 = db.Column(db.Boolean, default=False)
    day22 = db.Column(db.Boolean, default=False)
    day23 = db.Column(db.Boolean, default=False)
    day24 = db.Column(db.Boolean, default=False)
    day25 = db.Column(db.Boolean, default=False)
    day26 = db.Column(db.Boolean, default=False)
    day27 = db.Column(db.Boolean, default=False)
    day28 = db.Column(db.Boolean, default=False)
    day29 = db.Column(db.Boolean, default=False)
    day30 = db.Column(db.Boolean, default=False)
    mobile_number = db.Column(db.String(15), nullable=False)
    mm_registered_names = db.Column(db.String(200), nullable=False)
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    gps_location_name = db.Column(db.String(300), default='')
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)

    def get_day(self, n):
        return getattr(self, f'day{n}', False) or False

    def set_day(self, n, val):
        if 1 <= n <= 30:
            setattr(self, f'day{n}', bool(val))

    def to_dict(self):
        d = {
            'id': self.id,
            'assessment_id': self.assessment_id,
            'participant_name': self.participant_name,
            'cadre': self.cadre,
            'district': self.district,
            'facility': self.facility,
            'registration_date': self.registration_date.strftime('%Y-%m-%d'),
            'mobile_number': self.mobile_number,
            'mm_registered_names': self.mm_registered_names,
            'latitude': self.latitude,
            'longitude': self.longitude,
            'gps_location_name': self.gps_location_name or '',
            'submitted_at': self.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
        }
        for i in range(1, 31):
            d[f'day{i}'] = self.get_day(i)
        return d


VALID_PHONE_RE = re.compile(r'^\+256[0-9]{9}$')

VALID_DISTRICTS = {
    'Abim District','Adjumani District','Agago District','Alebtong District','Amolatar District',
    'Amudat District','Amuria District','Amuru District','Apac District','Arua City District',
    'Arua District','Budaka District','Bududa District','Bugiri District','Bugweri District',
    'Buhweju District','Buikwe District','Bukedea District','Bukomansimbi District','Bukwo District',
    'Bulambuli District','Buliisa District','Bundibugyo District','Bunyangabu District',
    'Bushenyi District','Busia District','Butaleja District','Butambala District','Butebo District',
    'Buvuma District','Buyende District','Dokolo District','Fort Portal City District',
    'Gomba District','Gulu City District','Gulu District','Hoima City District','Hoima District',
    'Ibanda District','Iganga District','Isingiro District','Jinja City District','Jinja District',
    'Kaabong District','Kabale District','Kabarole District','Kaberamaido District','Kagadi District',
    'Kakumiro District','Kalaki District','Kalangala District','Kaliro District','Kalungu District',
    'Kampala District','Kamuli District','Kamwenge District','Kanungu District','Kapchorwa District',
    'Kapelebyong District','Karenga District','Kasese District','Kassanda District',
    'Katakwi District','Kayunga District','Kazo District','Kibaale District','Kiboga District',
    'Kibuku District','Kikuube District','Kiruhura District','Kiryandongo District','Kisoro District',
    'Kitagwenda District','Kitgum District','Koboko District','Kole District','Kotido District',
    'Kumi District','Kwania District','Kween District','Kyankwanzi District','Kyegegwa District',
    'Kyenjojo District','Kyotera District','Lamwo District','Lira City District','Lira District',
    'Luuka District','Luwero District','Lwengo District','Lyantonde District','Madi-Okollo District',
    'Manafwa District','Maracha District','Masaka City District','Masaka District','Masindi District',
    'Mayuge District','Mbale City District','Mbale District','Mbarara City District',
    'Mbarara District','Mitooma District','Mityana District','Moroto District','Moyo District',
    'Mpigi District','Mubende District','Mukono District','Nabilatuk District',
    'Nakapiripirit District','Nakaseke District','Nakasongola District','Namayingo District',
    'Namisindwa District','Namutumba District','Napak District','Nebbi District','Ngora District',
    'Ntoroko District','Ntungamo District','Nwoya District','Obongi District','Omoro District',
    'Otuke District','Oyam District','Pader District','Pakwach District','Pallisa District',
    'Rakai District','Rubanda District','Rubirizi District','Rukiga District','Rukungiri District',
    'Rwampara District','Sembabule District','Serere District','Sheema District','Sironko District',
    'Soroti City District','Soroti District','Terego District','Tororo District','Wakiso District',
    'Yumbe District','Zombo District'
}


def validate_participant(p):
    """Server-side validation. Returns error string or None."""
    name = (p.get('participant_name') or '').strip()
    if not name or len(name) < 2:
        return 'Participant name is required (min 2 characters)'
    if not (p.get('cadre') or '').strip():
        return 'Cadre is required'
    district = (p.get('district') or '').strip()
    if not district:
        return 'District is required'
    if district not in VALID_DISTRICTS:
        return f'Invalid district: {district}'
    if not (p.get('facility') or '').strip():
        return 'Facility is required'
    phone = (p.get('mobile_number') or '').strip()
    if not VALID_PHONE_RE.match(phone):
        return f'Invalid phone number format: {phone}. Expected +256XXXXXXXXX'
    if not (p.get('mm_registered_names') or '').strip():
        return 'MoMo registered names required'
    if not p.get('registration_date'):
        return 'Registration date is required'
    return None


with app.app_context():
    db.create_all()
    # Migrate: add new columns if they don't exist
    if DATABASE_URL.startswith('postgresql'):
        try:
            for day_num in range(6, 31):
                col = f'day{day_num}'
                db.session.execute(db.text(
                    f"ALTER TABLE registration ADD COLUMN IF NOT EXISTS {col} BOOLEAN DEFAULT FALSE"
                ))
            # GPS columns
            db.session.execute(db.text(
                "ALTER TABLE registration ADD COLUMN IF NOT EXISTS latitude FLOAT"))
            db.session.execute(db.text(
                "ALTER TABLE registration ADD COLUMN IF NOT EXISTS longitude FLOAT"))
            # GPS location name
            db.session.execute(db.text(
                "ALTER TABLE registration ADD COLUMN IF NOT EXISTS gps_location_name VARCHAR(300) DEFAULT ''"))
            # Lock column
            db.session.execute(db.text(
                "ALTER TABLE assessment ADD COLUMN IF NOT EXISTS is_locked BOOLEAN DEFAULT FALSE"))
            # Bank detail participant name
            db.session.execute(db.text(
                "ALTER TABLE bank_detail ADD COLUMN IF NOT EXISTS participant_name VARCHAR(200) DEFAULT ''"))
            # Bank detail day columns
            for d in range(1, 31):
                col = f'day{d}'
                try:
                    db.session.execute(db.text(
                        f"ALTER TABLE bank_detail ADD COLUMN IF NOT EXISTS {col} BOOLEAN DEFAULT FALSE"))
                except Exception:
                    pass
            db.session.commit()
        except Exception:
            db.session.rollback()


# ── Auth ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


# ── Chart helpers ──────────────────────────────────────────────────

def bucket_stats(stats, top_n=10):
    """Take a list of (label, count) tuples and bucket beyond top_n into 'Others'."""
    if len(stats) <= top_n:
        return stats
    top = stats[:top_n]
    others_count = sum(s[1] for s in stats[top_n:])
    return list(top) + [('Others', others_count)]


# ── Excel builder ───────────────────────────────────────────────────

def build_excel(registrations, campaign_days, sheet_title="Registration & Attendance"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill(start_color="2B5097", end_color="2B5097", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    med_border = Border(
        left=Side(style='medium', color='000000'), right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'), bottom=Side(style='medium', color='000000'))
    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    checked_font = Font(name="Calibri", size=14, bold=True, color="008000")
    unchecked_font = Font(name="Calibri", size=14, color="AAAAAA")

    headers = ['No.', "Participant\u2019s Name", 'Cadre', 'Duty Station (Facility)', 'District',
               'Mobile Number Registered', 'Names Registered on Mobile Money']
    for day in range(1, campaign_days + 1):
        headers.append(f'Day {day}')
    headers.append('Registration Date')
    headers.append('GPS Location')
    headers.append('GPS Coordinates')
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = med_border
    ws.row_dimensions[1].height = 38

    for idx, reg in enumerate(registrations, start=1):
        row = [idx, reg.participant_name, reg.cadre, reg.facility, reg.district,
               reg.mobile_number, reg.mm_registered_names]
        for day in range(1, campaign_days + 1):
            row.append('\u2611' if getattr(reg, f'day{day}', False) else '\u2610')
        row.append(reg.registration_date.strftime('%Y-%m-%d') if reg.registration_date else '')
        row.append(reg.gps_location_name or '')
        coords = f'{reg.latitude:.4f}, {reg.longitude:.4f}' if reg.latitude and reg.longitude else ''
        row.append(coords)
        ws.append(row)

        row_num = idx + 1
        for col_num, cell in enumerate(ws[row_num], start=1):
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = cell_alignment
            if col_num == 1:
                cell.alignment = center_alignment
            if 7 < col_num <= 7 + campaign_days:
                cell.alignment = center_alignment
                cell.font = checked_font if cell.value == '\u2611' else unchecked_font
        ws.row_dimensions[row_num].height = 22

    base_widths = [5, 25, 16, 21, 16, 21, 30]
    all_widths = base_widths + [8] * campaign_days + [14, 35, 22]
    for i, w in enumerate(all_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    return wb


# ── Participant routes ──────────────────────────────────────────────

@app.route('/')
def index():
    """Show assessment selector for participants"""
    assessments = Assessment.query.filter_by(is_active=True).order_by(Assessment.created_at.desc()).all()
    return render_template('select_assessment.html', assessments=assessments)


@app.route('/join', methods=['POST'])
def join_assessment():
    """Participant enters assessment PIN to access form"""
    client_ip = request.remote_addr or '0.0.0.0'
    if _is_rate_limited(_failed_pin_attempts, client_ip):
        flash('Too many failed PIN attempts. Please try again in 15 minutes.', 'error')
        return redirect(url_for('index'))

    assessment_id = request.form.get('assessment_id')
    pin = request.form.get('pin', '').strip()

    if not assessment_id:
        flash('Please select an assessment.', 'error')
        return redirect(url_for('index'))

    assessment = Assessment.query.get(assessment_id)
    if not assessment or not assessment.is_active:
        flash('Assessment not found or inactive.', 'error')
        return redirect(url_for('index'))

    if assessment.pin != pin:
        _record_failure(_failed_pin_attempts, client_ip)
        flash('Incorrect PIN. Please get the correct PIN from your manager.', 'error')
        return redirect(url_for('index'))

    # Clear failed attempts on success
    _failed_pin_attempts.pop(client_ip, None)
    # Store in session
    session['participant_assessment_id'] = assessment.id
    return redirect(url_for('participant_menu', assessment_id=assessment.id))


@app.route('/menu/<int:assessment_id>')
def participant_menu(assessment_id):
    """Show menu with Field Activity and Bank Details options"""
    if session.get('participant_assessment_id') != assessment_id:
        flash('Please select and enter PIN for your assessment.', 'error')
        return redirect(url_for('index'))
    assessment = Assessment.query.get_or_404(assessment_id)
    if not assessment.is_active:
        flash('This assessment is no longer active.', 'error')
        return redirect(url_for('index'))
    return render_template('participant_menu.html', assessment=assessment)


@app.route('/register/<int:assessment_id>')
def registration_form(assessment_id):
    """Show registration form for a specific assessment"""
    # Verify session
    if session.get('participant_assessment_id') != assessment_id:
        flash('Please select and enter PIN for your assessment.', 'error')
        return redirect(url_for('index'))

    assessment = Assessment.query.get_or_404(assessment_id)
    if not assessment.is_active:
        flash('This assessment is no longer active.', 'error')
        return redirect(url_for('index'))

    return render_template('registration.html',
                         assessment=assessment,
                         campaign_days=assessment.campaign_days,
                         activity_name=assessment.name,
                         activity_dates=assessment.dates_label,
                         start_date=assessment.start_date.strftime('%Y-%m-%d') if assessment.start_date else '',
                         end_date=assessment.end_date.strftime('%Y-%m-%d') if assessment.end_date else '')


@app.route('/submit/bulk', methods=['POST'])
def submit_bulk_registration():
    try:
        data = request.get_json()
        participants = data.get('participants', [])
        assessment_id = data.get('assessment_id')
        gps_lat = data.get('latitude')
        gps_lng = data.get('longitude')
        gps_loc_name = data.get('gps_location_name', '')

        if not participants:
            return jsonify({'success': False, 'error': 'No participants provided'}), 400
        if not assessment_id:
            return jsonify({'success': False, 'error': 'No assessment specified'}), 400

        assessment = Assessment.query.get(assessment_id)
        if not assessment:
            return jsonify({'success': False, 'error': 'Assessment not found'}), 404
        if not assessment.is_active:
            return jsonify({'success': False, 'error': 'This assessment is no longer active'}), 403
        if assessment.is_locked:
            return jsonify({'success': False, 'error': 'This assessment is locked. No more submissions are allowed.'}), 403

        # Server-side validation
        for p in participants:
            err = validate_participant(p)
            if err:
                return jsonify({'success': False, 'error': err}), 400

        facility = participants[0].get('facility', 'Unknown')
        registrations = []

        for p in participants:
            reg_date = datetime.strptime(p.get('registration_date'), '%Y-%m-%d').date()
            registration = Registration(
                assessment_id=assessment_id,
                participant_name=p.get('participant_name', '').strip().title(),
                cadre=p.get('cadre', '').strip(),
                district=p.get('district', '').strip(),
                facility=p.get('facility', '').strip(),
                registration_date=reg_date,
                mobile_number=p.get('mobile_number', '').strip(),
                mm_registered_names=p.get('mm_registered_names', '').strip().title(),
                latitude=gps_lat,
                longitude=gps_lng,
                gps_location_name=gps_loc_name
            )
            for day_num in range(1, 31):
                registration.set_day(day_num, p.get(f'day{day_num}', False))
            db.session.add(registration)
            registrations.append(registration)

        db.session.commit()

        for reg in registrations:
            log_audit(assessment_id, 'create', 'registration', reg.id,
                      f'Participant: {reg.participant_name}, Facility: {reg.facility}')

        return jsonify({
            'success': True, 'facility': facility,
            'count': len(registrations),
            'data': [r.to_dict() for r in registrations]
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/download/facility/<int:assessment_id>/<facility_name>')
def download_facility_data(assessment_id, facility_name):
    try:
        assessment = Assessment.query.get_or_404(assessment_id)
        registrations = Registration.query.filter_by(
            assessment_id=assessment_id, facility=facility_name
        ).order_by(Registration.submitted_at.desc()).all()

        if not registrations:
            return jsonify({'success': False, 'error': 'No data found'}), 404

        wb = build_excel(registrations, assessment.campaign_days, "Facility Data")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = facility_name.replace(' ', '_').replace('/', '_')
        filename = f'CHAI_{safe_name}_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── Bank Details routes ─────────────────────────────────────────────

@app.route('/bank/<int:assessment_id>')
def bank_form(assessment_id):
    """Show bank details form for a specific assessment"""
    if session.get('participant_assessment_id') != assessment_id:
        flash('Please select and enter PIN for your assessment.', 'error')
        return redirect(url_for('index'))
    assessment = Assessment.query.get_or_404(assessment_id)
    if not assessment.is_active:
        flash('This assessment is no longer active.', 'error')
        return redirect(url_for('index'))
    return render_template('bank_form.html', assessment=assessment,
                         campaign_days=assessment.campaign_days)


@app.route('/submit/bank', methods=['POST'])
def submit_bank_details():
    try:
        data = request.get_json()
        members = data.get('members', [])
        assessment_id = data.get('assessment_id')
        if not members:
            return jsonify({'success': False, 'error': 'No members provided'}), 400
        if not assessment_id:
            return jsonify({'success': False, 'error': 'No assessment specified'}), 400
        assessment = Assessment.query.get(assessment_id)
        if not assessment:
            return jsonify({'success': False, 'error': 'Assessment not found'}), 404
        if not assessment.is_active:
            return jsonify({'success': False, 'error': 'This assessment is no longer active'}), 403
        if assessment.is_locked:
            return jsonify({'success': False, 'error': 'This assessment is locked. No more submissions allowed.'}), 403
        saved = []
        for m in members:
            bd = BankDetail(
                assessment_id=assessment_id,
                participant_name=(m.get('participant_name') or '').strip().title(),
                account_name=(m.get('account_name') or '').strip().title(),
                designation=(m.get('designation') or '').strip(),
                bank_name=(m.get('bank_name') or '').strip(),
                account_number=(m.get('account_number') or '').strip(),
                branch=(m.get('branch') or '').strip()
            )
            for d in range(1, 31):
                setattr(bd, f'day{d}', bool(m.get(f'day{d}', False)))
            if not bd.participant_name or not bd.account_name or not bd.bank_name or not bd.account_number:
                return jsonify({'success': False, 'error': 'Account name, bank name, and account number are required'}), 400
            db.session.add(bd)
            saved.append(bd)
        db.session.commit()
        for bd in saved:
            log_audit(assessment_id, 'create', 'bank_detail', bd.id, f'Account: {bd.account_name}, Bank: {bd.bank_name}')
        return jsonify({'success': True, 'count': len(saved)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/download/bank/<int:assessment_id>')
def download_bank_participant(assessment_id):
    """Participant-facing bank Excel download (no login required)"""
    assessment = Assessment.query.get_or_404(assessment_id)
    bank_details = BankDetail.query.filter_by(assessment_id=assessment_id).order_by(BankDetail.submitted_at.desc()).all()
    if not bank_details:
        return jsonify({'success': False, 'error': 'No data found'}), 404
    wb = build_bank_excel(bank_details, campaign_days=assessment.campaign_days)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = assessment.name.replace(' ', '_')
    filename = f'CHAI_Payment_{safe_name}_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)


def build_bank_excel(bank_details, sheet_title="Payment Tracker", campaign_days=5):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill(start_color="2B5097", end_color="2B5097", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    gold_fill = PatternFill(start_color="D4A843", end_color="D4A843", fill_type="solid")
    gold_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    med_border = Border(
        left=Side(style='medium', color='000000'), right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'), bottom=Side(style='medium', color='000000'))
    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    checked_font = Font(name="Calibri", size=14, bold=True, color="008000")
    unchecked_font = Font(name="Calibri", size=14, color="AAAAAA")

    # Base columns: No, Name, AccName, Desig, Bank, AccNo, Branch = 7 cols
    # Then day columns (campaign_days)
    # Then perdiem columns (10): Unit,Nights,Sub, Unit,Nights,Sub,SDA, Transport,Total,Comments
    base_cols = 7
    day_start = base_cols + 1  # col 8
    perdiem_start = day_start + campaign_days
    total_cols = perdiem_start + 9  # 10 perdiem cols

    # Row 1: PAYMENT DETAILS banner
    last_col_letter = openpyxl.utils.get_column_letter(total_cols)
    ws.merge_cells(f'A1:{last_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = "PAYMENT DETAILS"
    title_cell.fill = header_fill
    title_cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = med_border
    ws.row_dimensions[1].height = 30

    # Row 2: spacer
    ws.row_dimensions[2].height = 8

    # Row 3: Group headers
    # Attendance days group
    if campaign_days > 0:
        day_start_letter = openpyxl.utils.get_column_letter(day_start)
        day_end_letter = openpyxl.utils.get_column_letter(day_start + campaign_days - 1)
        ws.merge_cells(f'{day_start_letter}3:{day_end_letter}3')
        att_cell = ws[f'{day_start_letter}3']
        att_cell.value = "Attendance"
        att_cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        att_cell.font = gold_font
        att_cell.alignment = header_alignment
        att_cell.border = med_border

    # Perdiem travel nights (3 cols)
    p1_start = openpyxl.utils.get_column_letter(perdiem_start)
    p1_end = openpyxl.utils.get_column_letter(perdiem_start + 2)
    ws.merge_cells(f'{p1_start}3:{p1_end}3')
    g1 = ws[f'{p1_start}3']
    g1.value = "Perdiem for travel night"
    g1.fill = gold_fill
    g1.font = gold_font
    g1.alignment = header_alignment
    g1.border = med_border

    # Perdiem field days (4 cols)
    p2_start = openpyxl.utils.get_column_letter(perdiem_start + 3)
    p2_end = openpyxl.utils.get_column_letter(perdiem_start + 6)
    ws.merge_cells(f'{p2_start}3:{p2_end}3')
    g2 = ws[f'{p2_start}3']
    g2.value = "Perdiem field days and SDA return day"
    g2.fill = gold_fill
    g2.font = gold_font
    g2.alignment = header_alignment
    g2.border = med_border

    # Style remaining row 3 cells
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=3, column=col)
        if not cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = med_border
    ws.row_dimensions[3].height = 34

    # Row 4: Sub headers
    sub_headers = ['No.', "Participant\u2019s Name", 'Bank Account Name', 'Designation', 'Bank Name',
                   'Account Number', 'Bank Branch']
    for d in range(1, campaign_days + 1):
        sub_headers.append(f'Day {d}')
    sub_headers += ['Unit cost', '# of Nights', 'Sub-Total',
                    'Unit cost', '# of Nights', 'Sub-Total', 'SDA',
                    'Transport refund', 'Total Cost', 'Comments']
    for i, h in enumerate(sub_headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = med_border
    ws.row_dimensions[4].height = 38

    # Data rows
    for idx, bd in enumerate(bank_details, start=1):
        row_num = idx + 4
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=bd.participant_name or '')
        ws.cell(row=row_num, column=3, value=bd.account_name)
        ws.cell(row=row_num, column=4, value=bd.designation)
        ws.cell(row=row_num, column=5, value=bd.bank_name)
        ws.cell(row=row_num, column=6, value=bd.account_number)
        ws.cell(row=row_num, column=7, value=bd.branch)
        # Day columns
        for d in range(1, campaign_days + 1):
            col = base_cols + d
            val = '\u2611' if getattr(bd, f'day{d}', False) else '\u2610'
            ws.cell(row=row_num, column=col, value=val)
        # Style all columns
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = cell_alignment
            if col == 1:
                cell.alignment = center_alignment
            # Day columns styling
            if day_start <= col < perdiem_start:
                cell.alignment = center_alignment
                cell.font = checked_font if cell.value == '\u2611' else unchecked_font
            # Perdiem columns
            if col >= perdiem_start:
                cell.fill = empty_fill
                cell.alignment = center_alignment
        ws.row_dimensions[row_num].height = 22

    # Column widths
    widths = [5, 24, 24, 16, 18, 16, 15]
    widths += [8] * campaign_days  # day cols
    widths += [10, 12, 11, 10, 12, 11, 9, 15, 13, 21]  # perdiem cols
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return wb


# ── Admin routes ────────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        client_ip = request.remote_addr or '0.0.0.0'
        if _is_rate_limited(_failed_login_attempts, client_ip):
            flash('Too many failed login attempts. Please try again in 15 minutes.', 'error')
            return render_template('admin_login.html')

        username = request.form.get('username')
        password = request.form.get('password')
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['admin_logged_in'] = True
            session['admin_user'] = username
            # Clear failed attempts on success
            _failed_login_attempts.pop(client_ip, None)
            flash('Login successful!', 'success')
            return redirect(url_for('admin_assessments'))
        else:
            _record_failure(_failed_login_attempts, client_ip)
            flash('Invalid credentials.', 'error')
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_user', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('admin_login'))


@app.route('/admin/assessments')
@login_required
def admin_assessments():
    """List all assessments for this manager"""
    assessments = Assessment.query.order_by(Assessment.created_at.desc()).all()
    # Get registration counts per assessment
    counts = {}
    for a in assessments:
        counts[a.id] = Registration.query.filter_by(assessment_id=a.id).count()
    return render_template('admin_assessments.html', assessments=assessments, counts=counts)


@app.route('/admin/assessments/create', methods=['POST'])
@login_required
def create_assessment():
    name = request.form.get('name', '').strip()
    start_date_str = request.form.get('start_date', '').strip()
    end_date_str = request.form.get('end_date', '').strip()
    campaign_days = request.form.get('campaign_days', '3')
    pin = request.form.get('pin', '').strip()

    if not name:
        flash('Assessment name is required.', 'error')
        return redirect(url_for('admin_assessments'))

    # Parse dates
    start_date = None
    end_date = None
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        if start_date and end_date and end_date < start_date:
            flash('End date cannot be before start date.', 'error')
            return redirect(url_for('admin_assessments'))
    except ValueError:
        flash('Invalid date format.', 'error')
        return redirect(url_for('admin_assessments'))

    try:
        campaign_days_int = int(campaign_days)
        if campaign_days_int < 1 or campaign_days_int > 30:
            raise ValueError
    except ValueError:
        flash('Campaign days must be between 1 and 30.', 'error')
        return redirect(url_for('admin_assessments'))

    if not pin or len(pin) < 3:
        flash('PIN must be at least 3 characters.', 'error')
        return redirect(url_for('admin_assessments'))

    existing = Assessment.query.filter_by(pin=pin, is_active=True).first()
    if existing:
        flash('This PIN is already in use by another active assessment. Choose a different one.', 'error')
        return redirect(url_for('admin_assessments'))

    assessment = Assessment(
        name=name,
        start_date=start_date,
        end_date=end_date,
        campaign_days=campaign_days_int,
        pin=pin,
        created_by=session.get('admin_user', 'admin')
    )
    db.session.add(assessment)
    db.session.commit()
    flash(f'Assessment "{name}" created! PIN: {pin}', 'success')
    return redirect(url_for('admin_assessments'))


@app.route('/admin/assessments/<int:assessment_id>/toggle', methods=['POST'])
@login_required
def toggle_assessment(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    assessment.is_active = not assessment.is_active
    db.session.commit()
    status = 'activated' if assessment.is_active else 'deactivated'
    flash(f'Assessment "{assessment.name}" {status}.', 'success')
    return redirect(url_for('admin_assessments'))


@app.route('/admin/assessments/<int:assessment_id>/delete', methods=['POST'])
@login_required
def delete_assessment(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    # Delete all registrations first
    Registration.query.filter_by(assessment_id=assessment_id).delete()
    db.session.delete(assessment)
    db.session.commit()
    flash(f'Assessment "{assessment.name}" and all its data deleted.', 'success')
    return redirect(url_for('admin_assessments'))


@app.route('/admin/dashboard/<int:assessment_id>')
@login_required
def admin_dashboard(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    campaign_days = assessment.campaign_days

    search = request.args.get('search', '')
    district = request.args.get('district', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = Registration.query.filter_by(assessment_id=assessment_id)
    if search:
        query = query.filter(db.or_(
            Registration.participant_name.ilike(f'%{search}%'),
            Registration.mobile_number.ilike(f'%{search}%'),
            Registration.facility.ilike(f'%{search}%')
        ))
    if district:
        query = query.filter(Registration.district == district)
    if date_from:
        try:
            query = query.filter(Registration.registration_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Registration.registration_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 50
    filtered_query = query.order_by(Registration.submitted_at.desc())
    filtered_total = filtered_query.count()
    total_pages = max(1, (filtered_total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    registrations = filtered_query.offset((page - 1) * per_page).limit(per_page).all()

    districts = [d[0] for d in db.session.query(Registration.district).filter_by(
        assessment_id=assessment_id).distinct().order_by(Registration.district).all()]
    total_count = Registration.query.filter_by(assessment_id=assessment_id).count()
    today = datetime.utcnow().date()
    today_count = Registration.query.filter_by(assessment_id=assessment_id).filter(
        db.func.date(Registration.submitted_at) == today).count()

    # Enrollment trend by date
    daily_registrations = db.session.query(
        db.func.date(Registration.submitted_at).label('date'),
        db.func.count(Registration.id).label('count')
    ).filter(Registration.assessment_id == assessment_id
    ).group_by(db.func.date(Registration.submitted_at)).order_by('date').all()

    # District breakdown
    district_stats = db.session.query(
        Registration.district, db.func.count(Registration.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).group_by(Registration.district).order_by(db.func.count(Registration.id).desc()).all()

    # Facility breakdown
    facility_stats = db.session.query(
        Registration.facility, db.func.count(Registration.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).group_by(Registration.facility).order_by(db.func.count(Registration.id).desc()).all()

    # Cadre breakdown
    cadre_stats = db.session.query(
        Registration.cadre, db.func.count(Registration.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).group_by(Registration.cadre).order_by(db.func.count(Registration.id).desc()).all()

    # Per-day attendance counts
    ordinals = ['','1st','2nd','3rd','4th','5th','6th','7th','8th','9th','10th',
                '11th','12th','13th','14th','15th','16th','17th','18th','19th','20th',
                '21st','22nd','23rd','24th','25th','26th','27th','28th','29th','30th']
    day_attendance = []
    all_regs = Registration.query.filter_by(assessment_id=assessment_id).all()
    for d in range(1, campaign_days + 1):
        attended = sum(1 for r in all_regs if r.get_day(d))
        day_attendance.append({'day': d, 'label': ordinals[d], 'count': attended, 'total': len(all_regs)})

    # Unique counts
    unique_districts = db.session.query(db.func.count(db.distinct(Registration.district))).filter_by(
        assessment_id=assessment_id).scalar() or 0
    unique_facilities = db.session.query(db.func.count(db.distinct(Registration.facility))).filter_by(
        assessment_id=assessment_id).scalar() or 0

    # Bank details count for this assessment
    bank_count = BankDetail.query.filter_by(assessment_id=assessment_id).count()

    return render_template('admin_dashboard.html',
        assessment=assessment,
        registrations=registrations,
        total_count=total_count,
        today_count=today_count,
        filtered_count=filtered_total,
        search=search, districts=districts,
        selected_district=district,
        date_from=date_from, date_to=date_to,
        daily_registrations=daily_registrations,
        district_stats=bucket_stats(district_stats),
        facility_stats=bucket_stats(facility_stats),
        cadre_stats=bucket_stats(cadre_stats),
        day_attendance=day_attendance,
        unique_districts=unique_districts,
        unique_facilities=unique_facilities,
        bank_count=bank_count,
        campaign_days=campaign_days,
        page=page, total_pages=total_pages, per_page=per_page)


@app.route('/admin/settings/<int:assessment_id>', methods=['GET', 'POST'])
@login_required
def admin_settings(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)

    if request.method == 'POST':
        name = request.form.get('activity_name', '').strip()
        start_date_str = request.form.get('start_date', '').strip()
        end_date_str = request.form.get('end_date', '').strip()
        campaign_days = request.form.get('campaign_days', '3')
        pin = request.form.get('pin', '').strip()

        if name:
            assessment.name = name
        try:
            if start_date_str:
                assessment.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                assessment.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
        if pin and len(pin) >= 3:
            # Check uniqueness
            existing = Assessment.query.filter(
                Assessment.pin == pin, Assessment.is_active == True,
                Assessment.id != assessment_id).first()
            if existing:
                flash('PIN already in use by another assessment.', 'error')
                return redirect(url_for('admin_settings', assessment_id=assessment_id))
            assessment.pin = pin

        try:
            cd = int(campaign_days)
            if 1 <= cd <= 30:
                assessment.campaign_days = cd
        except ValueError:
            pass

        db.session.commit()
        flash('Settings updated!', 'success')
        return redirect(url_for('admin_settings', assessment_id=assessment_id))

    return render_template('admin_settings.html', assessment=assessment,
                         campaign_days=assessment.campaign_days,
                         activity_name=assessment.name)


@app.route('/admin/download/excel/<int:assessment_id>')
@login_required
def download_excel(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)

    district = request.args.get('district', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = Registration.query.filter_by(assessment_id=assessment_id)
    if district:
        query = query.filter(Registration.district == district)
    if date_from:
        try:
            query = query.filter(Registration.registration_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Registration.registration_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    registrations = query.order_by(Registration.submitted_at.desc()).all()
    wb = build_excel(registrations, assessment.campaign_days)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = assessment.name.replace(' ', '_')
    filename = f'CHAI_{safe_name}_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)


@app.route('/admin/delete/<int:assessment_id>/<int:reg_id>', methods=['POST'])
@login_required
def delete_registration(assessment_id, reg_id):
    registration = Registration.query.get_or_404(reg_id)
    assessment = Assessment.query.get_or_404(assessment_id)
    if assessment.is_locked:
        flash('Assessment is locked. Cannot delete.', 'error')
        return redirect(url_for('admin_dashboard', assessment_id=assessment_id))
    name = registration.participant_name
    db.session.delete(registration)
    db.session.commit()
    log_audit(assessment_id, 'delete', 'registration', reg_id, f'Deleted: {name}')
    flash('Registration deleted.', 'success')
    return redirect(url_for('admin_dashboard', assessment_id=assessment_id))


@app.route('/admin/clear-all/<int:assessment_id>', methods=['POST'])
@login_required
def clear_all(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    if assessment.is_locked:
        flash('Assessment is locked. Cannot clear data.', 'error')
        return redirect(url_for('admin_dashboard', assessment_id=assessment_id))
    count = Registration.query.filter_by(assessment_id=assessment_id).count()
    Registration.query.filter_by(assessment_id=assessment_id).delete()
    db.session.commit()
    log_audit(assessment_id, 'clear', 'registration', details=f'Cleared {count} registrations')
    flash('All registrations cleared.', 'success')
    return redirect(url_for('admin_dashboard', assessment_id=assessment_id))


@app.route('/admin/bank/<int:assessment_id>')
@login_required
def admin_bank_details(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)

    search = request.args.get('search', '')
    bank_filter = request.args.get('bank', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = BankDetail.query.filter_by(assessment_id=assessment_id)
    if search:
        query = query.filter(db.or_(
            BankDetail.participant_name.ilike(f'%{search}%'),
            BankDetail.account_name.ilike(f'%{search}%'),
            BankDetail.account_number.ilike(f'%{search}%'),
            BankDetail.designation.ilike(f'%{search}%'),
            BankDetail.branch.ilike(f'%{search}%')
        ))
    if bank_filter:
        query = query.filter(BankDetail.bank_name == bank_filter)
    if date_from:
        try:
            query = query.filter(BankDetail.submitted_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(BankDetail.submitted_at <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    # Pagination
    bank_page = request.args.get('page', 1, type=int)
    bank_per_page = 50
    bank_filtered_query = query.order_by(BankDetail.submitted_at.desc())
    bank_filtered_total = bank_filtered_query.count()
    bank_total_pages = max(1, (bank_filtered_total + bank_per_page - 1) // bank_per_page)
    bank_page = max(1, min(bank_page, bank_total_pages))
    bank_details = bank_filtered_query.offset((bank_page - 1) * bank_per_page).limit(bank_per_page).all()

    total_count = BankDetail.query.filter_by(assessment_id=assessment_id).count()
    banks = [b[0] for b in db.session.query(BankDetail.bank_name).filter_by(
        assessment_id=assessment_id).distinct().order_by(BankDetail.bank_name).all()]

    # Bank analytics
    bank_distribution = db.session.query(
        BankDetail.bank_name, db.func.count(BankDetail.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).group_by(BankDetail.bank_name).order_by(db.func.count(BankDetail.id).desc()).all()

    designation_stats = db.session.query(
        BankDetail.designation, db.func.count(BankDetail.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).filter(BankDetail.designation != ''
    ).group_by(BankDetail.designation).order_by(db.func.count(BankDetail.id).desc()).all()

    branch_stats = db.session.query(
        BankDetail.branch, db.func.count(BankDetail.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).filter(BankDetail.branch != ''
    ).group_by(BankDetail.branch).order_by(db.func.count(BankDetail.id).desc()).all()

    daily_submissions = db.session.query(
        db.func.date(BankDetail.submitted_at).label('date'),
        db.func.count(BankDetail.id).label('count')
    ).filter(BankDetail.assessment_id == assessment_id
    ).group_by(db.func.date(BankDetail.submitted_at)).order_by('date').all()

    unique_branches = db.session.query(db.func.count(db.distinct(BankDetail.branch))).filter_by(
        assessment_id=assessment_id).filter(BankDetail.branch != '').scalar() or 0

    return render_template('admin_bank.html',
        assessment=assessment, bank_details=bank_details,
        total_count=total_count, filtered_count=bank_filtered_total,
        search=search, banks=banks, selected_bank=bank_filter,
        date_from=date_from, date_to=date_to,
        bank_distribution=bucket_stats(bank_distribution),
        designation_stats=bucket_stats(designation_stats),
        branch_stats=bucket_stats(branch_stats),
        daily_submissions=daily_submissions,
        unique_branches=unique_branches,
        page=bank_page, total_pages=bank_total_pages, per_page=bank_per_page,
        campaign_days=assessment.campaign_days)


@app.route('/admin/bank/delete/<int:assessment_id>/<int:bd_id>', methods=['POST'])
@login_required
def delete_bank_detail(assessment_id, bd_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    if assessment.is_locked:
        flash('Assessment is locked. Cannot delete.', 'error')
        return redirect(url_for('admin_bank_details', assessment_id=assessment_id))
    bd = BankDetail.query.get_or_404(bd_id)
    name = bd.account_name
    db.session.delete(bd)
    db.session.commit()
    log_audit(assessment_id, 'delete', 'bank_detail', bd_id, f'Deleted: {name}')
    flash('Bank detail deleted.', 'success')
    return redirect(url_for('admin_bank_details', assessment_id=assessment_id))


@app.route('/admin/bank/clear/<int:assessment_id>', methods=['POST'])
@login_required
def clear_bank_details(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    if assessment.is_locked:
        flash('Assessment is locked. Cannot clear data.', 'error')
        return redirect(url_for('admin_bank_details', assessment_id=assessment_id))
    count = BankDetail.query.filter_by(assessment_id=assessment_id).count()
    BankDetail.query.filter_by(assessment_id=assessment_id).delete()
    db.session.commit()
    log_audit(assessment_id, 'clear', 'bank_detail', details=f'Cleared {count} bank details')
    flash('All bank details cleared.', 'success')
    return redirect(url_for('admin_bank_details', assessment_id=assessment_id))


@app.route('/admin/bank/download/<int:assessment_id>')
@login_required
def download_bank_excel(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)

    bank_filter = request.args.get('bank', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = BankDetail.query.filter_by(assessment_id=assessment_id)
    if bank_filter:
        query = query.filter(BankDetail.bank_name == bank_filter)
    if date_from:
        try:
            query = query.filter(BankDetail.submitted_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(BankDetail.submitted_at <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    bank_details = query.order_by(BankDetail.submitted_at.desc()).all()
    wb = build_bank_excel(bank_details, campaign_days=assessment.campaign_days)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = assessment.name.replace(' ', '_')
    filename = f'CHAI_Payment_{safe_name}_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)


# ── Edit routes ────────────────────────────────────────────────────

@app.route('/admin/edit/<int:assessment_id>/<int:reg_id>', methods=['GET', 'POST'])
@login_required
def edit_registration(assessment_id, reg_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    reg = Registration.query.get_or_404(reg_id)

    if assessment.is_locked:
        flash('Assessment is locked. Cannot edit.', 'error')
        return redirect(url_for('admin_dashboard', assessment_id=assessment_id))

    if request.method == 'POST':
        old_name = reg.participant_name
        reg.participant_name = request.form.get('participant_name', reg.participant_name)
        reg.cadre = request.form.get('cadre', reg.cadre)
        reg.district = request.form.get('district', reg.district)
        reg.facility = request.form.get('facility', reg.facility)
        reg.mobile_number = request.form.get('mobile_number', reg.mobile_number)
        reg.mm_registered_names = request.form.get('mm_registered_names', reg.mm_registered_names)

        for day_num in range(1, assessment.campaign_days + 1):
            reg.set_day(day_num, request.form.get(f'day{day_num}') == 'on')

        db.session.commit()
        log_audit(assessment_id, 'edit', 'registration', reg_id, f'Edited: {old_name}')
        flash('Registration updated.', 'success')
        return redirect(url_for('admin_dashboard', assessment_id=assessment_id))

    return render_template('edit_registration.html', assessment=assessment, reg=reg)


@app.route('/admin/bank/edit/<int:assessment_id>/<int:bd_id>', methods=['GET', 'POST'])
@login_required
def edit_bank_detail(assessment_id, bd_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    bd = BankDetail.query.get_or_404(bd_id)

    if assessment.is_locked:
        flash('Assessment is locked. Cannot edit.', 'error')
        return redirect(url_for('admin_bank_details', assessment_id=assessment_id))

    if request.method == 'POST':
        old_name = bd.account_name
        bd.participant_name = (request.form.get('participant_name') or bd.participant_name or '').strip().title()
        bd.account_name = (request.form.get('account_name') or bd.account_name).strip().title()
        bd.designation = request.form.get('designation', bd.designation)
        bd.bank_name = request.form.get('bank_name', bd.bank_name)
        bd.account_number = request.form.get('account_number', bd.account_number)
        bd.branch = request.form.get('branch', bd.branch)

        db.session.commit()
        log_audit(assessment_id, 'edit', 'bank_detail', bd_id, f'Edited: {old_name}')
        flash('Bank detail updated.', 'success')
        return redirect(url_for('admin_bank_details', assessment_id=assessment_id))

    return render_template('edit_bank.html', assessment=assessment, bd=bd)


# ── API ─────────────────────────────────────────────────────────────

@app.route('/api/assessments')
def api_assessments():
    assessments = Assessment.query.filter_by(is_active=True).all()
    return jsonify([a.to_dict() for a in assessments])


# ── Duplicate Detection ───────────────────────────────────────────

def normalize_name(name):
    """Normalize name for comparison: lowercase, strip, collapse spaces"""
    return ' '.join((name or '').lower().strip().split())

def names_similar(n1, n2):
    """Check if two names are similar (same words in any order, or substring)"""
    w1 = set(normalize_name(n1).split())
    w2 = set(normalize_name(n2).split())
    if not w1 or not w2:
        return False
    # Same words in any order
    if w1 == w2:
        return True
    # One is subset of the other (catches "John Doe" vs "John Doe Okello")
    if w1.issubset(w2) or w2.issubset(w1):
        return True
    # High overlap (e.g. 2 out of 3 words match)
    overlap = len(w1 & w2)
    total = max(len(w1), len(w2))
    if total >= 2 and overlap >= total - 1:
        return True
    return False


@app.route('/admin/location-alerts/<int:assessment_id>')
@login_required
def admin_location_alerts(assessment_id):
    """Detect submissions where GPS location doesn't match selected district"""
    regs = Registration.query.filter_by(assessment_id=assessment_id).filter(
        Registration.latitude.isnot(None),
        Registration.gps_location_name != '',
        Registration.gps_location_name.isnot(None)
    ).all()

    mismatches = []
    for r in regs:
        dist_name = r.district.replace(' District', '').replace(' City', '').lower()
        loc_lower = (r.gps_location_name or '').lower()
        if dist_name and loc_lower and dist_name not in loc_lower:
            mismatches.append({
                'id': r.id,
                'name': r.participant_name,
                'selected_district': r.district,
                'gps_location': r.gps_location_name[:80],
                'facility': r.facility,
                'latitude': r.latitude,
                'longitude': r.longitude,
                'date': r.registration_date.strftime('%Y-%m-%d') if r.registration_date else ''
            })

    return jsonify({'mismatches': mismatches, 'count': len(mismatches)})


@app.route('/admin/duplicates/<int:assessment_id>')
@login_required
def admin_duplicates(assessment_id):
    """Detect duplicate phone numbers and similar names"""
    # Phone duplicates
    phone_dupes = db.session.query(
        Registration.mobile_number,
        db.func.count(Registration.id).label('cnt')
    ).filter_by(assessment_id=assessment_id
    ).group_by(Registration.mobile_number
    ).having(db.func.count(Registration.id) > 1).all()

    phone_results = []
    for phone, cnt in phone_dupes:
        regs = Registration.query.filter_by(
            assessment_id=assessment_id, mobile_number=phone
        ).all()
        phone_results.append({
            'mobile_number': phone,
            'count': cnt,
            'registrations': [{'id': r.id, 'name': r.participant_name,
                              'facility': r.facility, 'district': r.district} for r in regs]
        })

    # Name duplicates (fuzzy) - group by facility to limit comparisons
    all_regs = Registration.query.filter_by(assessment_id=assessment_id).all()
    name_results = []
    seen_pairs = set()
    for i, r1 in enumerate(all_regs):
        for r2 in all_regs[i+1:]:
            pair_key = tuple(sorted([r1.id, r2.id]))
            if pair_key in seen_pairs:
                continue
            if normalize_name(r1.participant_name) == normalize_name(r2.participant_name):
                match_type = 'exact'
            elif names_similar(r1.participant_name, r2.participant_name):
                match_type = 'similar'
            else:
                continue
            seen_pairs.add(pair_key)
            name_results.append({
                'match_type': match_type,
                'reg1': {'id': r1.id, 'name': r1.participant_name,
                         'facility': r1.facility, 'phone': r1.mobile_number},
                'reg2': {'id': r2.id, 'name': r2.participant_name,
                         'facility': r2.facility, 'phone': r2.mobile_number}
            })
            # Limit to 50 name matches to avoid overload
            if len(name_results) >= 50:
                break
        if len(name_results) >= 50:
            break

    # Cross-assessment phone duplicates
    cross_results = []
    current_phones = db.session.query(Registration.mobile_number).filter_by(
        assessment_id=assessment_id).distinct().all()
    current_phone_set = {p[0] for p in current_phones}
    if current_phone_set:
        other_regs = Registration.query.filter(
            Registration.assessment_id != assessment_id,
            Registration.mobile_number.in_(current_phone_set)
        ).all()
        # Group by phone
        cross_map = {}
        for r in other_regs:
            if r.mobile_number not in cross_map:
                cross_map[r.mobile_number] = []
            cross_map[r.mobile_number].append(r)
        for phone, regs_list in cross_map.items():
            # Get the assessment names
            assessment_ids_set = {r.assessment_id for r in regs_list}
            assessments_info = Assessment.query.filter(Assessment.id.in_(assessment_ids_set)).all()
            assess_names = {a.id: a.name for a in assessments_info}
            cross_results.append({
                'mobile_number': phone,
                'current_assessment': assessment.name,
                'other_assessments': [
                    {'assessment_name': assess_names.get(r.assessment_id, 'Unknown'),
                     'name': r.participant_name, 'facility': r.facility}
                    for r in regs_list[:10]  # limit per phone
                ]
            })

    return jsonify({
        'phone_duplicates': phone_results,
        'name_duplicates': name_results,
        'cross_assessment_duplicates': cross_results
    })


# ── Bulk Attendance Edit ──────────────────────────────────────────

@app.route('/admin/bulk-attendance/<int:assessment_id>', methods=['POST'])
@login_required
def bulk_attendance(assessment_id):
    """Bulk update attendance days for multiple registrations"""
    try:
        assessment = Assessment.query.get_or_404(assessment_id)
        if assessment.is_locked:
            return jsonify({'success': False, 'error': 'Assessment is locked'}), 403

        data = request.get_json()
        reg_ids = data.get('reg_ids', [])
        days = data.get('days', {})
        if not reg_ids or not days:
            return jsonify({'success': False, 'error': 'No registrations or days specified'}), 400

        updated = 0
        for reg_id in reg_ids:
            reg = Registration.query.filter_by(id=reg_id, assessment_id=assessment_id).first()
            if not reg:
                continue
            for day_str, value in days.items():
                day_num = int(day_str)
                if 1 <= day_num <= 30:
                    setattr(reg, f'day{day_num}', bool(value))
            updated += 1

        db.session.commit()
        day_list = ', '.join([f'Day {k}={"Present" if v else "Absent"}' for k, v in days.items()])
        log_audit(assessment_id, 'bulk_edit', 'registration',
                  details=f'Bulk updated {updated} registrations: {day_list}')
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


# ── Lock/Unlock Assessment ────────────────────────────────────────

@app.route('/admin/assessments/<int:assessment_id>/lock', methods=['POST'])
@login_required
def toggle_lock(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    assessment.is_locked = not assessment.is_locked
    db.session.commit()
    action = 'lock' if assessment.is_locked else 'unlock'
    log_audit(assessment_id, action, 'assessment', assessment_id,
              f'Assessment {"locked" if assessment.is_locked else "unlocked"}')
    flash(f'Assessment {"locked" if assessment.is_locked else "unlocked"}.', 'success')
    return redirect(url_for('admin_settings', assessment_id=assessment_id))


# ── Audit Log Viewer ──────────────────────────────────────────────

@app.route('/admin/audit/<int:assessment_id>')
@login_required
def admin_audit_log(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    logs = AuditLog.query.filter_by(assessment_id=assessment_id).order_by(
        AuditLog.performed_at.desc()).limit(200).all()
    return render_template('admin_audit.html', assessment=assessment, logs=logs)


@app.route('/admin/summary-pdf/<int:assessment_id>')
@login_required
def admin_summary_pdf(assessment_id):
    """Generate a printable HTML summary page for an assessment"""
    assessment = Assessment.query.get_or_404(assessment_id)

    total_count = Registration.query.filter_by(assessment_id=assessment_id).count()
    unique_districts = db.session.query(db.func.count(db.distinct(Registration.district))).filter_by(
        assessment_id=assessment_id).scalar() or 0
    unique_facilities = db.session.query(db.func.count(db.distinct(Registration.facility))).filter_by(
        assessment_id=assessment_id).scalar() or 0

    district_stats = db.session.query(
        Registration.district, db.func.count(Registration.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).group_by(Registration.district).order_by(db.func.count(Registration.id).desc()).limit(10).all()

    facility_stats = db.session.query(
        Registration.facility, db.func.count(Registration.id).label('count')
    ).filter_by(assessment_id=assessment_id
    ).group_by(Registration.facility).order_by(db.func.count(Registration.id).desc()).limit(10).all()

    date_range = ''
    if assessment.start_date and assessment.end_date:
        date_range = f"{assessment.start_date.strftime('%d %b %Y')} to {assessment.end_date.strftime('%d %b %Y')}"
    elif assessment.start_date:
        date_range = f"From {assessment.start_date.strftime('%d %b %Y')}"

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Summary - {assessment.name}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'DM Sans',sans-serif;background:#fff;color:#1F2937;padding:40px}}
.header{{text-align:center;margin-bottom:30px;border-bottom:3px solid #0F5B5C;padding-bottom:20px}}
.header h1{{color:#0F5B5C;font-size:24px;margin-bottom:4px}}
.header p{{font-size:13px;color:#6B7280}}
.stats{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:30px}}
.stat{{text-align:center;padding:16px;border:2px solid #E5E7EB;border-radius:8px}}
.stat .num{{font-size:28px;font-weight:700;color:#0F5B5C}}
.stat .label{{font-size:11px;color:#6B7280;margin-top:4px}}
table{{width:100%;border-collapse:collapse;margin-bottom:24px;font-size:13px}}
th,td{{padding:8px 12px;text-align:left;border:1px solid #E5E7EB}}
thead{{background:#0F5B5C;color:white}}
th{{font-weight:600;font-size:11px;text-transform:uppercase}}
tbody tr:nth-child(even){{background:#F9FAFB}}
h2{{color:#0A3D3E;font-size:16px;margin-bottom:12px}}
.print-btn{{background:#0F5B5C;color:white;border:none;padding:10px 24px;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;font-family:inherit;margin-bottom:20px}}
.print-btn:hover{{background:#178182}}
@media print{{
 .print-btn{{display:none}}
 body{{padding:20px}}
}}
</style>
</head>
<body>
<button class="print-btn" onclick="window.print()">Print / Save as PDF</button>
<div class="header">
 <h1>{assessment.name}</h1>
 <p>{date_range} &middot; CHAI Uganda Field Operations</p>
 <p style="margin-top:4px;font-size:11px;color:#9CA3AF">Generated on {datetime.utcnow().strftime("%d %b %Y %H:%M")} UTC</p>
</div>
<div class="stats">
 <div class="stat"><div class="num">{total_count}</div><div class="label">Total Participants</div></div>
 <div class="stat"><div class="num">{unique_districts}</div><div class="label">Districts Covered</div></div>
 <div class="stat"><div class="num">{unique_facilities}</div><div class="label">Facilities Reached</div></div>
 <div class="stat"><div class="num">{assessment.campaign_days}</div><div class="label">Campaign Days</div></div>
</div>
<h2>Top Districts</h2>
<table><thead><tr><th>No.</th><th>District</th><th>Participants</th></tr></thead><tbody>'''
    for i, (dist, cnt) in enumerate(district_stats, 1):
        html += f'<tr><td>{i}</td><td>{dist}</td><td>{cnt}</td></tr>'
    html += '''</tbody></table>
<h2>Top Facilities</h2>
<table><thead><tr><th>No.</th><th>Facility</th><th>Participants</th></tr></thead><tbody>'''
    for i, (fac, cnt) in enumerate(facility_stats, 1):
        html += f'<tr><td>{i}</td><td>{fac}</td><td>{cnt}</td></tr>'
    html += '''</tbody></table>
</body></html>'''
    return html


@app.route('/healthz')
def healthz():
    return jsonify(status='ok')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
